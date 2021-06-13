import os, sys
import pip
import openpyxl


def import_or_install(package):
    try:
        __import__(package)
    except ImportError:
        print("Trying to install " + package + " ...")
        pip.main(['install', package])

def clear_old_excel_rows(book):
    sheetnames = book.sheetnames
    for sheetname in sheetnames:
        sheet = book[sheetname]
        header_size = 2 # First row will hold disclaimer, title in 2nd row
        active_rows = len(sheet['A'])
        if active_rows < header_size:
            print("Info: sheet \""+sheetname+"\" is new / fresh! So, not clearing old data!")
            return
        sheet.delete_rows(header_size+1, active_rows)        


def open_excel_out_file(file):
    if os.path.isfile(file):
        try:
            book = openpyxl.load_workbook(file, read_only=False)
        except:
            print(Fore.YELLOW +"Error: Please close " + file + " and retry!")
            print(Style.RESET_ALL)
            return None
    else:
        print("Creating \"" + file + "\"")
        book = openpyxl.Workbook()
    return book

def is_hex(s):
    try:
        int(s, 16)
        return True
    except ValueError:
        return False